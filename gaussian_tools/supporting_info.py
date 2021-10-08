#!/usr/bin/env python3

'''
Author: Inigo Iribarren
26-03-2020

##################
## DESCRIPTIONS ##
##################

Script to extract the supporting information for all the files in a folder and 
generates a excel file with:
    - SCF energy
    - Free energy
    - N imaginary frequencies
    - Coordinates

This script needs 2 files for each molecule:
  name.log for the COORDINATES AND FREQUENCIES
  name_SP.log fot the SCF ENERGY

###########
## USAGE ##
###########

    supporting_info 

'''
# SUPPORTING INFORMATION SCRIPT

# Imports
import glob
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from my_functions import last_find, first_find

# Code for the elements 
code = {"1" : "H", "2" : "He", "3" : "Li", "4" : "Be", "5" : "B", \
"6"  : "C", "7"  : "N", "8"  : "O",  "9" : "F", "10" : "Ne", \
"11" : "Na" , "12" : "Mg" , "13" : "Al" , "14" : "Si" , "15" : "P", \
"16" : "S"  , "17" : "Cl" , "18" : "Ar" , "19" : "K"  , "20" : "Ca", \
"21" : "Sc" , "22" : "Ti" , "23" : "V"  , "24" : "Cr" , "25" : "Mn", \
"26" : "Fe" , "27" : "Co" , "28" : "Ni" , "29" : "Cu" , "30" : "Zn", \
"31" : "Ga" , "32" : "Ge" , "33" : "As" , "34" : "Se" , "35" : "Br", \
"36" : "Kr" , "37" : "Rb" , "38" : "Sr" , "39" : "Y"  , "40" : "Zr", \
"41" : "Nb" , "42" : "Mo" , "43" : "Tc" , "44" : "Ru" , "45" : "Rh", \
"46" : "Pd" , "47" : "Ag" , "48" : "Cd" , "49" : "In" , "50" : "Sn", \
"51" : "Sb" , "52" : "Te" , "53" : "I"  , "54" : "Xe" , "55" : "Cs", \
"56" : "Ba" , "57" : "La" , "58" : "Ce" , "59" : "Pr" , "60" : "Nd", \
"61" : "Pm" , "62" : "Sm" , "63" : "Eu" , "64" : "Gd" , "65" : "Tb", \
"66" : "Dy" , "67" : "Ho" , "68" : "Er" , "69" : "Tm" , "70" : "Yb", \
"71" : "Lu" , "72" : "Hf" , "73" : "Ta" , "74" : "W"  , "75" : "Re", \
"76" : "Os" , "77" : "Ir" , "78" : "Pt" , "79" : "Au" , "80" : "Hg", \
"81" : "Tl" , "82" : "Pb" , "83" : "Bi" , "84" : "Po" , "85" : "At", \
"86" : "Rn" , "87" : "Fr" , "88" : "Ra" , "89" : "Ac" , "90" : "Th", \
"91" : "Pa" , "92" : "U"  , "93" : "Np" , "94" : "Pu" , "95" : "Am", \
"96" : "Cm" , "97" : "Bk" , "98" : "Cf" , "99" : "Es" ,"100" : "Fm", \
"101": "Md" ,"102" : "No" ,"103" : "Lr" ,"104" : "Rf" ,"105" : "Db", \
"106": "Sg" ,"107" : "Bh" ,"108" : "Hs" ,"109" : "Mt" ,"110" : "Ds", \
"111": "Rg" ,"112" : "Uub","113" : "Uut","114" : "Uuq","115" : "Uup", \
"116": "Uuh","117" : "Uus","118" : "Uuo"}

# Start the code for ALL THE .LOG FILES    

for file in glob.glob("*.log"):
    name = file.split(".")[0]
    if name.split("_")[-1] != "SP":
        # Name the different files
        dz = name+'.log'
        tz = name+'_SP.log'

        # Load the text files:
        logfile = open(dz, 'r')
        text_dz = logfile.readlines()
        logfile.close()

        logfile = open(tz, 'r')
        text_tz = logfile.readlines()
        logfile.close()

        # Last coordinates:
        coord = text_dz

        coord = coord[last_find(coord, 'Coordinates')+3:]
        coord = coord[:first_find(coord, '----')]

        # Translation of the atomic numbers to the atomic symbol and create the matrix with the coordinates
        m_coord = []
        for line in coord:
            column = line.split()
            symbol = code.get(column[1], 'X')
 
            m_coord.append([symbol, column[3], column[4], column[5]])
    
        #print(m_coord)

        # Look for imaginary frequencies (and save them just in case)
        freq = text_dz
        freqs = []
        img_freq = 0

        for line in freq:
            if (line.find('Frequencies')) > -1:
                for f in line.split()[2:]:
                    if float(f) < 0:
                        img_freq += 1
                    freqs.append(float(f))

        #print(freqs)
        #print(img_freq)

        # Energy from pVTZ file
        scf = text_tz

        scf_energy = scf[last_find(scf,'SCF Done')].split()[4]
        #print(scf_energy)

        # Free energy from pVDZ
        dG_text = text_dz
        try:
            dG = dG_text[last_find(dG_text, 'Sum of electronic and thermal Free Energies')].split()[7]
        except:
            dG = "?\t"

        # Save everything in excel

        filename="supporting_information.xlsx"
        if os.path.isfile(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook() 

        sheet = wb.create_sheet(name)

        sheet["A1"] = "Compound"
        sheet.merge_cells('B1:E1')
        sheet["B1"] = "Coordinates"
        sheet["B2"] = "Atom"
        sheet["C2"] = "x"
        sheet["D2"] = "y"
        sheet["E2"] = "z"

        sheet["A3"] = name
        sheet["A4"] = "Img. freq. = " + str(img_freq)
        sheet["A5"] = "SCF Energy (a.u.) = " + str(scf_energy)
        sheet["A6"] = "Free Energy (a.u.) = " + str(dG)

        if (len(m_coord) >= 6):
            sheet.merge_cells('A7:A'+str(2+len(m_coord)))

        for i in range(0,len(m_coord)):
            sheet['B'+str(3 + i)] = m_coord[i][0]
            sheet['C'+str(3 + i)] = m_coord[i][1]
            sheet['D'+str(3 + i)] = m_coord[i][2]
            sheet['E'+str(3 + i)] = m_coord[i][3]

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for i in range(1,3+len(m_coord)):
            for j in range(1,6):
                sheet.cell(row=i, column=j).border = thin_border
        print(name, " done")


        # Save the excel file
        wb.save(filename)
